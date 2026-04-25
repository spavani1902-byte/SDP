from docx import Document
from docx.table import Table
from docx.text.paragraph import Paragraph
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
import re
import os


def iter_block_items(parent):
    for child in parent.element.body:
        if child.tag.endswith('}p'):
            yield Paragraph(child, parent)
        elif child.tag.endswith('}tbl'):
            yield Table(child, parent)


def convert_word_to_excel(file_path):

    doc = Document(file_path)
    wb = Workbook()
    ws = wb.active
    row_num = 1

    bold_font = Font(bold=True)
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_align_top = Alignment(horizontal="left", vertical="top", wrap_text=True)

    border_style = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    all_paragraphs = [p for p in doc.paragraphs if p.text.strip()]
    header_limit = 6

    for i in range(min(len(all_paragraphs), header_limit)):
        text = all_paragraphs[i].text.strip()

        ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=5)
        cell = ws.cell(row=row_num, column=1, value=text)
        cell.font = bold_font
        cell.alignment = center_align
        row_num += 1

    row_num += 1

    current_part = None
    instruction = None
    blooms_text = None

    for block in iter_block_items(doc):

        if isinstance(block, Paragraph):
            text = block.text.strip()
            if not text:
                continue

            if any(text == p.text.strip() for p in all_paragraphs[:header_limit]):
                continue

            if "Bloom" in text or "L-1" in text:
                blooms_text = text
                continue

            if text.upper().startswith("PART"):
                current_part = text
                continue

            if re.search(r'\d+\s*[xX*]\s*\d+', text):
                instruction = text
                continue

        elif isinstance(block, Table):

            if current_part:
                ws.cell(row=row_num, column=1, value=current_part).font = bold_font
                row_num += 1

            if instruction:
                ws.cell(row=row_num, column=1, value=instruction)
                row_num += 1

            if current_part == "PART-A" and blooms_text:
                ws.cell(row=row_num, column=1, value=blooms_text).font = bold_font
                row_num += 1

            first_row = True

            for row in block.rows:
                raw_cells = [cell.text.strip() for cell in row.cells]

                clean_cells = []
                if raw_cells:
                    clean_cells.append(raw_cells[0])
                    for i in range(1, len(raw_cells)):
                        if raw_cells[i] != raw_cells[i-1]:
                            clean_cells.append(raw_cells[i])

                for col_idx, value in enumerate(clean_cells, start=1):
                    c = ws.cell(row=row_num, column=col_idx, value=value)
                    c.border = border_style

                    if first_row:
                        c.font = bold_font
                        c.alignment = center_align
                    else:
                        c.alignment = left_align_top

                row_num += 1
                first_row = False

            current_part = None
            instruction = None
            row_num += 2

    column_widths = {'A': 6, 'B': 55, 'C': 10, 'D': 10, 'E': 10}
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width

    # SAVE FILE
    output_path = file_path.replace(".docx", "_converted.xlsx")
    wb.save(output_path)

    return output_path