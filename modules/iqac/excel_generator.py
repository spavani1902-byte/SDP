from openpyxl import load_workbook


# =========================
# SAFE WRITE (MERGED CELLS FIX)
# =========================
def get_top_left(ws, cell):
    for merged in ws.merged_cells.ranges:
        if cell in merged:
            return merged.start_cell.coordinate
    return cell


def write_safe(ws, cell, value):
    real_cell = get_top_left(ws, cell)
    ws[real_cell] = value


# =========================
# MAIN FUNCTION
# =========================
def generate_iqac_excel(df, template_path, output_path):

    wb = load_workbook(template_path)
    ws = wb.active

    # 🔥 FIX COLUMN NAMES
    df.columns = df.columns.str.strip()
    df["Marks"] = df["Marks"].astype(int)

    # =========================
    # BLOOM LEVEL TABLE
    # =========================

    levels = ["L1","L2","L3","L4","L5","L6"]
    level_rows = [9, 10, 11, 12, 13, 14]

    for level, row in zip(levels, level_rows):

        partA = df[(df["BL"] == level) & (df["Part"] == "A")]
        partB = df[(df["BL"] == level) & (df["Part"] == "B")]

        qnos_A = ",".join(partA["QNo"])
        marks_A = partA["Marks"].sum() if not partA.empty else 0

        qnos_B = ",".join(partB["QNo"])
        marks_B = partB["Marks"].sum() if not partB.empty else 0

        write_safe(ws, f"E{row}", qnos_A)
        write_safe(ws, f"F{row}", marks_A)
        write_safe(ws, f"G{row}", qnos_B)
        write_safe(ws, f"H{row}", marks_B)
        write_safe(ws, f"I{row}", marks_A + marks_B)

    # =========================
    # CO TABLE
    # =========================

    cos = ["CO1","CO2","CO3","CO4","CO5","CO6"]
    co_rows = [21, 22, 23, 24, 25, 26]

    for co, row in zip(cos, co_rows):

        partA = df[(df["CO"] == co) & (df["Part"] == "A")]
        partB = df[(df["CO"] == co) & (df["Part"] == "B")]

        qnos_A = ",".join(partA["QNo"])
        marks_A = partA["Marks"].sum() if not partA.empty else 0

        qnos_B = ",".join(partB["QNo"])
        marks_B = partB["Marks"].sum() if not partB.empty else 0

        write_safe(ws, f"E{row}", qnos_A)
        write_safe(ws, f"F{row}", marks_A)
        write_safe(ws, f"G{row}", qnos_B)
        write_safe(ws, f"H{row}", marks_B)
        write_safe(ws, f"I{row}", marks_A + marks_B)

    wb.save(output_path)
    return output_path