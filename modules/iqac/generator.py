from openpyxl import load_workbook
from collections import defaultdict
import shutil

def sort_questions(q_list):
    def key(q):
        num = int(''.join(filter(str.isdigit, q)))
        sub = ''.join(filter(str.isalpha, q))
        return (num, sub)
    return sorted(q_list, key=key)


def update_mapping(output_path, parsed):

    template_path = "modules/iqac/template.xlsx"
    shutil.copy(template_path, output_path)

    wb = load_workbook(output_path)
    print(wb.sheetnames)
    ws = wb["Exam Paper"]

    bloom_partA = defaultdict(list)
    bloom_partB = defaultdict(list)
    bloom_marksA = defaultdict(int)
    bloom_marksB = defaultdict(int)

    co_partA = defaultdict(list)
    co_partB = defaultdict(list)
    co_marksA = defaultdict(int)
    co_marksB = defaultdict(int)

    for item in parsed:

        q = item["q"]
        level = item["level"]
        co = item["co"]
        marks = item["marks"]
        part = item["part"]

        if part == "A":
            bloom_partA[level].append(q)
            bloom_marksA[level] += marks
            co_partA[co].append(q)
            co_marksA[co] += marks
        else:
            bloom_partB[level].append(q)
            bloom_marksB[level] += marks
            co_partB[co].append(q)
            co_marksB[co] += marks

    bloom_row_map = {"L1":10,"L2":11,"L3":12,"L4":13,"L5":14,"L6":15}

    for level, row in bloom_row_map.items():
        ws[f"E{row}"] = ",".join(sort_questions(bloom_partA[level]))
        ws[f"F{row}"] = bloom_marksA[level]
        ws[f"G{row}"] = ",".join(sort_questions(bloom_partB[level]))
        ws[f"H{row}"] = bloom_marksB[level]
        ws[f"I{row}"] = bloom_marksA[level] + bloom_marksB[level]

    co_row_map = {"CO1":21,"CO2":22,"CO3":23,"CO4":24,"CO5":25,"CO6":26}

    for co, row in co_row_map.items():
        ws[f"E{row}"] = ",".join(sort_questions(co_partA[co]))
        ws[f"F{row}"] = co_marksA[co]
        ws[f"G{row}"] = ",".join(sort_questions(co_partB[co]))
        ws[f"H{row}"] = co_marksB[co]
        ws[f"I{row}"] = co_marksA[co] + co_marksB[co]

    wb.save(output_path)

    print("✅ IQAC Excel Generated:", output_path)

    return output_path